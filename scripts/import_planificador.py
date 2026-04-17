#!/usr/bin/env python3
"""
Importa hojas "Planificador horario semana N" desde un Excel al formato JSON
del Planificador web (compatible con el botón Importar de la app).

USO
───
  python scripts/import_planificador.py
  python scripts/import_planificador.py --entrada mi_archivo.xlsx
  python scripts/import_planificador.py --salida backup.json
  python scripts/import_planificador.py --debug        # muestra JSON intermedio
  python scripts/import_planificador.py --semana 16    # solo esa semana

ESTRUCTURA ESPERADA DEL EXCEL (por hoja)
─────────────────────────────────────────
  Fila 1   : "Planificador horario semana 16 - Nombre"
  Fila ~2  : "Semana de: 16-04-2026"  ← fecha inicio de la semana
  Filas 4-7: Encabezados de secciones: "Objetivos semanales" | "Elementos pendientes" | "Emergencias"
  Filas 8-13: Contenido de esas tres secciones (una ítem por celda)
  Fila ~20 : Números de día (16, 17, 18 …) — el script los detecta dinámicamente
  Fila ~21 : Nombre del mes bajo cada día
  Filas 22-41: Tareas. Cada día ocupa 2 col: [estado ✓/✗/vacío, descripción]
  Filas 51-59: Notas de la semana (col de "Objetivos semanales")

MAPEO AL MODELO DE LA APP
──────────────────────────
  Objetivos semanales  →  Goal (scope='mensual', category='Trabajo')
  Elementos pendientes →  Task (priority='media') en el primer día de la semana
  Emergencias          →  Task (priority='alta')  en el primer día de la semana
  Tareas por día       →  Task dentro de DailyPlan del día correspondiente
    ✓ verde            →  status='completada'
    ✗ rojo             →  status='cancelada'
    vacío              →  status='pendiente'
  Notas                →  NoteBlock[] en DailyPlan.note del primer día
"""

import argparse
import json
import re
import sys
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

# ── month name → number ───────────────────────────────────────────────────────
MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4,
    "mayo": 5, "junio": 6, "julio": 7, "agosto": 8,
    "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
}

DIAS_NOMBRE = ["lunes", "martes", "miércoles", "jueves", "viernes", "sábado", "domingo"]

# Characters / strings that mean "done" ✓
DONE_CHARS = {"✓", "✔", "v", "si", "sí", "1", "ok", "x̌"}
# Characters / strings that mean "cancelled" ✗
CANCEL_CHARS = {"✗", "✘", "×", "x", "no", "0", "n"}

SCHEMA_PREFIX = "planner:v1"
GOAL_COLORS = ["#6366f1", "#22c55e", "#f59e0b", "#ef4444", "#8b5cf6", "#06b6d4"]


# ── helpers ───────────────────────────────────────────────────────────────────

def new_id() -> str:
    return str(uuid.uuid4())


def now_iso() -> str:
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.000Z")


def cell_text(cell) -> str:
    """Return stripped string value of a cell, or ''."""
    if cell is None or cell.value is None:
        return ""
    return str(cell.value).strip()


def parse_status(raw: str) -> str:
    """Map a status cell value to 'completada' | 'cancelada' | 'pendiente'."""
    s = raw.strip().lower()
    if s in DONE_CHARS:
        return "completada"
    if s in CANCEL_CHARS:
        return "cancelada"
    return "pendiente"


def parse_semana_date(text: str) -> date | None:
    """
    Extract a date from strings like:
      "Semana de: 16-04-2026"
      "16/04/2026"
      "16-04-2026"
    Returns a date object or None.
    """
    # Try to find DD-MM-YYYY or DD/MM/YYYY anywhere in the string
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{4})", text)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


def week_number_from_title(title: str) -> int | None:
    m = re.search(r"semana\s+(\d+)", title, re.IGNORECASE)
    return int(m.group(1)) if m else None


# ── sheet parser ──────────────────────────────────────────────────────────────

def find_row_with_date(ws, max_row: int = 10) -> tuple[date | None, int]:
    """
    Scan the first *max_row* rows looking for a cell containing a date string
    like "Semana de: 16-04-2026".
    Returns (date, row_index_1based) or (None, -1).
    """
    for r in range(1, max_row + 1):
        for cell in ws[r]:
            txt = cell_text(cell)
            if txt:
                d = parse_semana_date(txt)
                if d:
                    return d, r
    return None, -1


def find_section_headers(ws, keyword_map: dict, search_rows: range) -> dict[str, int]:
    """
    Given a dict of {label: keyword}, scan *search_rows* for cells whose text
    contains each keyword. Returns {label: col_index_1based}.
    """
    found = {}
    for r in search_rows:
        for cell in ws[r]:
            txt = cell_text(cell).lower()
            for label, kw in keyword_map.items():
                if label not in found and kw in txt:
                    found[label] = cell.column
        if len(found) == len(keyword_map):
            break
    return found


def find_day_header_row(ws, search_rows: range) -> int:
    """
    Find the row that contains the most integer day-numbers (1-31).
    Returns 1-based row index or -1.
    """
    best_row, best_count = -1, 0
    for r in search_rows:
        count = 0
        for cell in ws[r]:
            v = cell.value
            if isinstance(v, (int, float)) and 1 <= int(v) <= 31:
                count += 1
        if count > best_count:
            best_count, best_row = count, r
    return best_row if best_count >= 2 else -1


def extract_day_columns(ws, day_row: int, start_date: date) -> list[dict]:
    """
    From the day-header row, build a list of:
      { "col": int, "fecha": date, "nombre": str }
    sorted by column.

    Strategy:
      1. Collect all columns in day_row that have an integer 1-31.
      2. For each, look in the next 1-3 rows for a month name to confirm the date.
      3. If no month found, infer date from start_date + offset.
    """
    days = []
    for cell in ws[day_row]:
        v = cell.value
        if not (isinstance(v, (int, float)) and 1 <= int(v) <= 31):
            continue
        day_num = int(v)
        col = cell.column

        # Search adjacent rows for month name
        month_num = None
        day_name = ""
        for delta in range(1, 4):
            r2 = day_row + delta
            if r2 > ws.max_row:
                break
            nearby = cell_text(ws.cell(row=r2, column=col)).lower()
            if not month_num and nearby in MESES:
                month_num = MESES[nearby]
            if not day_name and nearby in {
                "lunes","martes","miércoles","miercoles",
                "jueves","viernes","sábado","sabado","domingo"
            }:
                day_name = nearby

        # Also check same row ±1 col for day name (sometimes it's inline)
        if not day_name:
            for dc in (-1, 1):
                c2 = col + dc
                if c2 < 1:
                    continue
                nearby = cell_text(ws.cell(row=day_row, column=c2)).lower()
                if nearby in {
                    "lunes","martes","miércoles","miercoles",
                    "jueves","viernes","sábado","sabado","domingo"
                }:
                    day_name = nearby
                    break

        year = start_date.year
        if month_num:
            try:
                d = date(year, month_num, day_num)
            except ValueError:
                d = start_date  # fallback
        else:
            # Infer: find the day within ±10 days of start_date
            d = None
            for offset in range(-3, 11):
                candidate = start_date + timedelta(days=offset)
                if candidate.day == day_num:
                    d = candidate
                    break
            if d is None:
                d = start_date

        # Derive day name from date if not found in the sheet
        if not day_name:
            day_name = DIAS_NOMBRE[d.weekday()]

        days.append({"col": col, "fecha": d, "nombre": day_name})

    # Sort by column
    days.sort(key=lambda x: x["col"])
    return days


def extract_tasks_for_day(ws, day_col: int, task_rows: range) -> list[dict]:
    """
    For a given day anchor column, extract tasks from the given rows.
    Layout per row: [status_col=day_col, description_col=day_col+1]
    """
    tasks = []
    for r in task_rows:
        status_cell = ws.cell(row=r, column=day_col)
        desc_cell   = ws.cell(row=r, column=day_col + 1)

        desc = cell_text(desc_cell)
        if not desc:
            continue  # empty task row

        raw_status = cell_text(status_cell)
        status = parse_status(raw_status)

        tasks.append({
            "texto":     desc,
            "hecho":     status == "completada",
            "cancelada": status == "cancelada",
            "status":    status,
        })
    return tasks


def extract_column_items(ws, col: int, rows: range) -> list[str]:
    """Collect non-empty text values from a single column across given rows."""
    items = []
    for r in rows:
        txt = cell_text(ws.cell(row=r, column=col))
        if txt:
            items.append(txt)
    return items


def parse_sheet(ws) -> dict | None:
    """
    Parse one sheet. Returns the intermediate dict or None if not parseable.
    """
    title = cell_text(ws.cell(row=1, column=1))

    # Accept sheets with merged-cell titles: scan row 1
    if not title:
        for cell in ws[1]:
            t = cell_text(cell)
            if t:
                title = t
                break

    week_num = week_number_from_title(title)
    if week_num is None:
        return None  # not a planificador sheet

    # ── exact start date ─────────────────────────────────────────────────────
    start_date, _date_row = find_row_with_date(ws, max_row=8)
    if start_date is None:
        # Fallback: try to derive from ISO week
        today = date.today()
        d = date.fromisocalendar(today.year, week_num, 1)
        start_date = d

    # ── section header columns (Objetivos, Elementos, Emergencias) ───────────
    sections = find_section_headers(ws, {
        "objetivos":  "objetivos",
        "pendientes": "pendientes",
        "emergencias": "emergencias",
    }, search_rows=range(3, 10))

    obj_col   = sections.get("objetivos", 3)       # default col C
    pend_col  = sections.get("pendientes")
    emerg_col = sections.get("emergencias")

    item_rows = range(8, 16)  # rows 8-15
    objetivos  = extract_column_items(ws, obj_col,   item_rows)
    pendientes = extract_column_items(ws, pend_col,  item_rows) if pend_col else []
    emergencias= extract_column_items(ws, emerg_col, item_rows) if emerg_col else []

    # ── day-header row (dynamic detection) ───────────────────────────────────
    day_row = find_day_header_row(ws, search_rows=range(15, 30))
    if day_row == -1:
        print(f"  ⚠  No se encontró fila de días en hoja '{ws.title}'")
        return None

    day_cols = extract_day_columns(ws, day_row, start_date)
    if not day_cols:
        print(f"  ⚠  No se detectaron columnas de días en hoja '{ws.title}'")
        return None

    # ── task rows: from day_row+2 to a sensible end ───────────────────────────
    task_start = day_row + 2
    task_end   = min(task_start + 25, ws.max_row)
    # Try to detect end of task block (large run of empty rows)
    task_rows = range(task_start, task_end + 1)

    days_data = []
    for day_info in day_cols:
        tareas = extract_tasks_for_day(ws, day_info["col"], task_rows)
        days_data.append({
            "fecha":    day_info["fecha"].strftime("%Y-%m-%d"),
            "dia":      day_info["nombre"],
            "tareas":   tareas,
        })

    # ── notes ─────────────────────────────────────────────────────────────────
    notes_start = task_end + 5
    notes_end   = min(notes_start + 15, ws.max_row)
    # Try to find a "Notas" header to anchor the search
    for r in range(task_end, min(task_end + 20, ws.max_row + 1)):
        for cell in ws[r]:
            if "nota" in cell_text(cell).lower():
                notes_start = r + 1
                notes_end   = min(r + 12, ws.max_row)
                break

    notas = extract_column_items(ws, obj_col, range(notes_start, notes_end + 1))

    return {
        "semana":      week_num,
        "año":         start_date.year,
        "fecha_inicio": start_date.strftime("%Y-%m-%d"),
        "objetivos":   objetivos,
        "pendientes":  pendientes,
        "emergencias": emergencias,
        "dias":        days_data,
        "notas":       notas,
    }


# ── Planner JSON builder ──────────────────────────────────────────────────────

def build_task(texto: str, status: str, due_date: str, priority: str = "media") -> dict:
    t = {
        "id":        new_id(),
        "title":     texto,
        "priority":  priority,
        "status":    status,
        "dueDate":   due_date,
        "tags":      [],
        "createdAt": now_iso(),
        "updatedAt": now_iso(),
    }
    if status == "completada":
        t["completedAt"] = now_iso()
    return t


def build_goal(title: str, year_month: str, color_idx: int) -> dict:
    year, month = map(int, year_month.split("-"))
    return {
        "id":          new_id(),
        "title":       title,
        "status":      "no_iniciada",
        "scope":       "mensual",
        "year":        year,
        "month":       month,
        "progress":    0,
        "category":    "Trabajo",
        "color":       GOAL_COLORS[color_idx % len(GOAL_COLORS)],
        "createdAt":   now_iso(),
        "updatedAt":   now_iso(),
    }


def build_note_block(text: str, order: int) -> dict:
    return {
        "id":      new_id(),
        "type":    "paragraph",
        "content": text,
        "order":   order,
    }


def week_to_planner_json(week: dict) -> dict:
    """Convert one parsed week dict → localStorage-compatible JSON fragment."""
    output = {}
    first_date = week["fecha_inicio"]
    year_month = first_date[:7]  # "YYYY-MM"

    # ── daily plans ──────────────────────────────────────────────────────────
    for i, day in enumerate(week["dias"]):
        date_str = day["fecha"]
        daily_key = f"{SCHEMA_PREFIX}:daily:{date_str}"

        tasks = [
            build_task(t["texto"], t["status"], date_str, "media")
            for t in day["tareas"]
        ]

        # Inject Elementos pendientes + Emergencias into first day
        if i == 0:
            for txt in week["pendientes"]:
                tasks.append(build_task(txt, "pendiente", date_str, "media"))
            for txt in week["emergencias"]:
                tasks.append(build_task(txt, "pendiente", date_str, "alta"))

        daily_plan: dict = {
            "date":         date_str,
            "tasks":        tasks,
            "habitEntries": [],
        }

        # Attach notas to the first day
        if i == 0 and week["notas"]:
            blocks = [build_note_block(n, j) for j, n in enumerate(week["notas"])]
            daily_plan["note"] = {
                "id":        new_id(),
                "scopeType": "daily",
                "scopeKey":  date_str,
                "blocks":    blocks,
                "updatedAt": now_iso(),
            }

        # Merge with existing entry if present (multiple weeks may share a month)
        if daily_key in output:
            output[daily_key]["tasks"].extend(tasks)
        else:
            output[daily_key] = daily_plan

    # ── monthly goals (Objetivos semanales) ──────────────────────────────────
    monthly_key = f"{SCHEMA_PREFIX}:monthly:{year_month}"
    goals = [
        build_goal(obj, year_month, idx)
        for idx, obj in enumerate(week["objetivos"])
    ]
    if goals:
        if monthly_key in output:
            output[monthly_key]["goals"].extend(goals)
        else:
            output[monthly_key] = {
                "yearMonth": year_month,
                "goals":     goals,
            }

    return output


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Importa 'Planificador horario semana N' → JSON para Planner web."
    )
    parser.add_argument("--entrada",  default="data/planificador_semanal.xlsx")
    parser.add_argument("--salida",   default="data/output_planificador.json")
    parser.add_argument("--debug",    action="store_true",
                        help="Guarda también el JSON intermedio en data/debug_semanas.json")
    parser.add_argument("--semana",   type=int, default=None,
                        help="Importar solo la semana N (ignora el resto)")
    args = parser.parse_args()

    entrada = Path(args.entrada)
    salida  = Path(args.salida)

    if not entrada.exists():
        print(f"Archivo no encontrado: {entrada}")
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print("Instalá openpyxl:  pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(entrada, data_only=True)

    # Filter sheets matching the expected pattern
    target_sheets = [
        s for s in wb.sheetnames
        if re.search(r"planificador\s+horario\s+semana", s, re.IGNORECASE)
    ]

    if not target_sheets:
        print(f"No se encontraron hojas 'Planificador horario semana N' en {entrada}")
        print(f"Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    print(f"Hojas encontradas: {len(target_sheets)}")

    parsed_weeks = []
    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        result = parse_sheet(ws)
        if result is None:
            print(f"  ✗  {sheet_name} — no procesada")
            continue
        if args.semana and result["semana"] != args.semana:
            continue

        n_tasks = sum(len(d["tareas"]) for d in result["dias"])
        print(f"  ✓  Semana {result['semana']:>2}  ({result['fecha_inicio']})  "
              f"— {n_tasks} tareas, {len(result['objetivos'])} objetivos, "
              f"{len(result['notas'])} notas")
        parsed_weeks.append(result)

    if not parsed_weeks:
        print("Nada para exportar.")
        sys.exit(1)

    # ── debug intermediate JSON ───────────────────────────────────────────────
    if args.debug:
        debug_path = salida.parent / "debug_semanas.json"
        debug_path.parent.mkdir(parents=True, exist_ok=True)
        with open(debug_path, "w", encoding="utf-8") as f:
            json.dump(parsed_weeks, f, ensure_ascii=False, indent=2, default=str)
        print(f"\nJSON intermedio: {debug_path}")

    # ── build final planner JSON ──────────────────────────────────────────────
    final: dict = {}
    for week in parsed_weeks:
        fragment = week_to_planner_json(week)
        for key, value in fragment.items():
            if key in final and isinstance(value, dict) and "tasks" in value:
                final[key]["tasks"].extend(value["tasks"])
            elif key in final and isinstance(value, dict) and "goals" in value:
                final[key]["goals"].extend(value["goals"])
            else:
                final[key] = value

    salida.parent.mkdir(parents=True, exist_ok=True)
    with open(salida, "w", encoding="utf-8") as f:
        json.dump(final, f, ensure_ascii=False, indent=2)

    total_tasks = sum(
        len(v.get("tasks", [])) for v in final.values() if isinstance(v, dict)
    )
    total_goals = sum(
        len(v.get("goals", [])) for v in final.values() if isinstance(v, dict)
    )
    print(f"\n✓  {salida}")
    print(f"   {len(parsed_weeks)} semana(s) · {total_tasks} tareas · {total_goals} objetivos")
    print(f"\nCómo importar:")
    print(f"  1. Planner → barra lateral → 'Importar'")
    print(f"  2. Seleccioná: {salida}")
    print(f"  3. Recargá la página (F5)")


if __name__ == "__main__":
    main()
